import json
import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox
import ttkbootstrap as tb
from ttkbootstrap.constants import *
from pathlib import Path
import warnings
warnings.filterwarnings("ignore", category=UserWarning)

try:
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, TwoCellAnchor, AnchorMarker
    from openpyxl.utils import coordinate_to_tuple
except ImportError as e:
    import traceback
    err_msg = traceback.format_exc()
    messagebox.showerror("匯入程式庫失敗", f"缺失必要組件，請回報以下資訊:\n{err_msg}")
    sys.exit(1)

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

try:
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    HAS_XDR = True
except ImportError:
    HAS_XDR = False

# ---------- 单位转换 ----------
def cm_to_px(cm_val):
    return int(cm_val * 37.795)

def cm_to_emu(cm_val):
    return int(cm_val * 360000)

def get_cell_value(wb, sheet_name, cell_ref):
    if isinstance(wb, xlrd.Book):
        sheet = wb.sheet_by_name(sheet_name)
        row, col = coordinate_to_tuple(cell_ref)
        return sheet.cell_value(row - 1, col - 1)
    else:
        ws = wb[sheet_name]
        return ws[cell_ref].value

CONFIG_FILE = "update_config.json"

def save_config(data):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def load_config():
    if not os.path.exists(CONFIG_FILE):
        return get_default_config()
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        cfg = json.load(f)
    if "data_source_path" in cfg and "data_sources" not in cfg:
        old_path = cfg.pop("data_source_path")
        cfg["data_sources"] = [{"alias": "預設資料來源", "path": old_path}]
        for m in cfg.get("data_mappings", []):
            if "source_alias" not in m:
                m["source_alias"] = "預設資料來源"
    for m in cfg.get("data_mappings", []):
        if "note" not in m:
            m["note"] = ""
    for m in cfg.get("image_mappings", []):
        if "position" not in m:
            m["position"] = "top-left"
            m["offset_x_cm"] = 0
            m["offset_y_cm"] = 0
        if "note" not in m:
            m["note"] = ""
    if "output_dir" not in cfg:
        cfg["output_dir"] = ""
    return cfg

def get_default_config():
    return {
        "data_sources": [],
        "template_path": "",
        "output_dir": "",
        "output_suffix": "_已更新",
        "data_mappings": [],
        "image_mappings": []
    }

# ========== 占位符工具函数 ==========
def add_placeholder(entry, placeholder_text):
    """为 Entry 添加占位符效果"""
    entry._placeholder_active = True
    entry._placeholder_text = placeholder_text
    entry.insert(0, placeholder_text)
    entry.config(foreground="gray")

    def on_focus_in(event):
        if entry._placeholder_active:
            entry.delete(0, tk.END)
            entry.config(foreground="black")
            entry._placeholder_active = False

    def on_focus_out(event):
        if not entry.get().strip():
            entry.delete(0, tk.END)
            entry.insert(0, placeholder_text)
            entry.config(foreground="gray")
            entry._placeholder_active = True

    entry.bind("<FocusIn>", on_focus_in)
    entry.bind("<FocusOut>", on_focus_out)

def get_entry_value(entry):
    """获取 Entry 的真实值，忽略占位符"""
    if hasattr(entry, "_placeholder_active") and entry._placeholder_active:
        return ""
    return entry.get().strip()

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("VSA_MBO_PBO 一鍵產生器")
        self.config = load_config()
        self.current_config_name = CONFIG_FILE
        self.style = self.root.style

        # ---------- 自定义按钮样式 ----------
        self.style.configure("CustomInfo.TButton", **self.style.configure("info.TButton"))
        self.style.map("CustomInfo.TButton", **self.style.map("info.TButton"))
        self.style.configure("CustomInfo.TButton", font=("", 10, "bold"))

        # ---------- 路径设置 ----------
        path_frame = tb.LabelFrame(root, text="範本與輸出設定", padding=10, bootstyle="info")
        path_frame.pack(fill=tk.X, padx=10, pady=5)

        self.tpl_path_var = tk.StringVar(value=self.config.get("template_path", ""))
        self.out_dir_var = tk.StringVar(value=self.config.get("output_dir", ""))
        self.suffix_var = tk.StringVar(value=self.config.get("output_suffix", "_已更新"))

        tb.Label(path_frame, text="範本檔案:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
        tb.Entry(path_frame, textvariable=self.tpl_path_var, width=70).grid(row=0, column=1, padx=5, pady=2)
        tb.Button(path_frame, text="瀏覽", command=self.browse_tpl, bootstyle="secondary-outline").grid(row=0, column=2, padx=5)

        tb.Label(path_frame, text="輸出資料夾:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
        tb.Entry(path_frame, textvariable=self.out_dir_var, width=70).grid(row=1, column=1, padx=5, pady=2)
        tb.Button(path_frame, text="瀏覽", command=self.browse_out_dir, bootstyle="secondary-outline").grid(row=1, column=2, padx=5)

        tb.Label(path_frame, text="輸出檔案後綴:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=2)
        tb.Entry(path_frame, textvariable=self.suffix_var, width=15).grid(row=2, column=1, sticky=tk.W, padx=5)

        tb.Label(path_frame, text="輸出資料夾留空則儲存至範本所在目錄", foreground="gray", font=("微软雅黑", 9)).grid(row=3, column=1, sticky=tk.W, padx=5)

        # ---------- 数据源管理 ----------
        ds_frame = tb.LabelFrame(root, text="資料來源管理", padding=10, bootstyle="info")
        ds_frame.pack(fill=tk.X, padx=10, pady=5)

        ds_container = tb.Frame(ds_frame)
        ds_container.pack(fill=tk.X, padx=5, pady=5)
        self.ds_tree = tb.Treeview(ds_container, columns=("alias", "path"), show="headings", height=3, bootstyle="primary")
        self.ds_tree.heading("alias", text="別名")
        self.ds_tree.heading("path", text="路徑")
        self.ds_tree.column("alias", width=150)
        self.ds_tree.column("path", width=600)
        vsb_ds = tb.Scrollbar(ds_container, orient=VERTICAL, command=self.ds_tree.yview, bootstyle="round")
        self.ds_tree.configure(yscrollcommand=vsb_ds.set)
        self.ds_tree.grid(row=0, column=0, sticky="nsew")
        vsb_ds.grid(row=0, column=1, sticky="ns")
        ds_container.grid_columnconfigure(0, weight=1)

        btn_ds = tb.Frame(ds_frame)
        btn_ds.pack(fill=tk.X, padx=5, pady=2)
        tb.Button(btn_ds, text="新增資料來源", command=self.add_datasource, bootstyle="primary-outline").pack(side=tk.LEFT, padx=5)
        tb.Button(btn_ds, text="編輯選取", command=self.edit_datasource, bootstyle="secondary-outline").pack(side=tk.LEFT, padx=5)
        tb.Button(btn_ds, text="刪除選取", command=self.delete_datasource, bootstyle="danger-outline").pack(side=tk.LEFT, padx=5)

        # ---------- 映射 Notebook ----------
        nb = tb.Notebook(root)
        nb.pack(side=tk.TOP, fill=tk.BOTH, expand=tk.YES, padx=10, pady=5)

        # ---- 数据映射页 ----
        data_frame = tb.Frame(nb)
        nb.add(data_frame, text="資料對應")

        data_container = tb.Frame(data_frame)
        data_container.pack(fill=tk.BOTH, expand=tk.YES, padx=5, pady=5)
        self.data_tree = tb.Treeview(data_container, columns=("source_alias", "source_cell", "target_cell", "note"), show="headings", bootstyle="primary")
        self.data_tree.heading("source_alias", text="資料來源")
        self.data_tree.heading("source_cell", text="來源儲存格")
        self.data_tree.heading("target_cell", text="目標儲存格")
        self.data_tree.heading("note", text="備註")
        self.data_tree.column("source_alias", width=120)
        self.data_tree.column("source_cell", width=160)
        self.data_tree.column("target_cell", width=160)
        self.data_tree.column("note", width=200)
        vsb_data = tb.Scrollbar(data_container, orient=VERTICAL, command=self.data_tree.yview, bootstyle="round")
        self.data_tree.configure(yscrollcommand=vsb_data.set)
        self.data_tree.grid(row=0, column=0, sticky="nsew")
        vsb_data.grid(row=0, column=1, sticky="ns")
        data_container.grid_rowconfigure(0, weight=1)
        data_container.grid_columnconfigure(0, weight=1)

        btn_data = tb.Frame(data_frame)
        btn_data.pack(fill=tk.X, padx=5, pady=2)
        tb.Button(btn_data, text="新增", command=self.add_data_mapping, bootstyle="primary-outline").pack(side=tk.LEFT, padx=5)
        tb.Button(btn_data, text="編輯選取", command=self.edit_data_mapping, bootstyle="secondary-outline").pack(side=tk.LEFT, padx=5)
        tb.Button(btn_data, text="複製選取", command=self.copy_data_mapping, bootstyle="info-outline").pack(side=tk.LEFT, padx=5)
        tb.Button(btn_data, text="刪除選取", command=lambda: self.delete_selected(self.data_tree, "data"), bootstyle="danger-outline").pack(side=tk.LEFT, padx=5)

        # ---- 图片映射页 ----
        img_frame = tb.Frame(nb)
        nb.add(img_frame, text="圖片對應")

        img_container = tb.Frame(img_frame)
        img_container.pack(fill=tk.BOTH, expand=tk.YES, padx=5, pady=5)
        self.img_tree = tb.Treeview(img_container, columns=("number", "folder", "target", "height", "width", "position", "note"), show="headings", bootstyle="primary")
        self.img_tree.heading("number", text="圖片編號")
        self.img_tree.heading("folder", text="圖片資料夾")
        self.img_tree.heading("target", text="目標儲存格")
        self.img_tree.heading("height", text="高度(cm)")
        self.img_tree.heading("width", text="寬度(cm)")
        self.img_tree.heading("position", text="位置")
        self.img_tree.heading("note", text="備註")
        self.img_tree.column("number", width=80)
        self.img_tree.column("folder", width=180)
        self.img_tree.column("target", width=100)
        self.img_tree.column("height", width=70)
        self.img_tree.column("width", width=70)
        self.img_tree.column("position", width=90)
        self.img_tree.column("note", width=180)
        vsb_img = tb.Scrollbar(img_container, orient=VERTICAL, command=self.img_tree.yview, bootstyle="round")
        self.img_tree.configure(yscrollcommand=vsb_img.set)
        self.img_tree.grid(row=0, column=0, sticky="nsew")
        vsb_img.grid(row=0, column=1, sticky="ns")
        img_container.grid_rowconfigure(0, weight=1)
        img_container.grid_columnconfigure(0, weight=1)

        btn_img = tb.Frame(img_frame)
        btn_img.pack(fill=tk.X, padx=5, pady=2)
        tb.Button(btn_img, text="新增", command=self.add_image_mapping, bootstyle="primary-outline").pack(side=tk.LEFT, padx=5)
        tb.Button(btn_img, text="編輯選取", command=self.edit_image_mapping, bootstyle="secondary-outline").pack(side=tk.LEFT, padx=5)
        tb.Button(btn_img, text="複製選取", command=self.copy_image_mapping, bootstyle="info-outline").pack(side=tk.LEFT, padx=5)
        tb.Button(btn_img, text="刪除選取", command=lambda: self.delete_selected(self.img_tree, "image"), bootstyle="danger-outline").pack(side=tk.LEFT, padx=5)

        # ---------- 控制按钮 ----------
        ctrl_frame = tb.Frame(root)
        ctrl_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=10)

        btn_container = tb.Frame(ctrl_frame)
        btn_container.pack(anchor=tk.CENTER)

        tb.Button(btn_container, text="⚡ 一鍵更新報告", command=self.run_update,
                  bootstyle="info", style="CustomInfo.TButton").pack(side=tk.LEFT, padx=10)
        tb.Button(btn_container, text="匯出設定", command=self.export_config,
                  bootstyle="secondary-outline").pack(side=tk.LEFT, padx=10)
        tb.Button(btn_container, text="匯入設定", command=self.import_config,
                  bootstyle="secondary-outline").pack(side=tk.LEFT, padx=10)

        self.refresh_datasource_tree()
        self.refresh_data_tree()
        self.refresh_image_tree()

    # ---------- 浏览 ----------
    def browse_tpl(self):
        path = filedialog.askopenfilename(filetypes=[("Excel檔案", "*.xlsx;*.xls")])
        if path:
            self.tpl_path_var.set(path)

    def browse_out_dir(self):
        path = filedialog.askdirectory()
        if path:
            self.out_dir_var.set(path)

    # ---------- 数据源管理 ----------
    def refresh_datasource_tree(self):
        for i in self.ds_tree.get_children():
            self.ds_tree.delete(i)
        for ds in self.config.get("data_sources", []):
            self.ds_tree.insert("", tk.END, values=(ds["alias"], ds["path"]))

    def add_datasource(self):
        self._datasource_dialog(None)

    def edit_datasource(self):
        selected = self.ds_tree.selection()
        if not selected:
            messagebox.showinfo("提示", "請先選擇一個資料來源")
            return
        idx = self.ds_tree.index(selected[0])
        item = self.config["data_sources"][idx]
        self._datasource_dialog(idx, item)

    def _datasource_dialog(self, edit_idx, item=None):
        popup = tb.Toplevel(self.root)
        popup.title("編輯資料來源" if item else "新增資料來源")
        tb.Label(popup, text="別名:").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        alias_entry = tb.Entry(popup, width=30)
        alias_entry.grid(row=0, column=1, padx=5, pady=5)
        add_placeholder(alias_entry, "非破壞性數據")
        tb.Label(popup, text="路徑:").grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        path_var = tk.StringVar()
        path_entry = tb.Entry(popup, textvariable=path_var, width=30)
        path_entry.grid(row=1, column=1, padx=5, pady=5)
        tb.Button(popup, text="瀏覽", command=lambda: path_var.set(filedialog.askopenfilename(filetypes=[("Excel檔案", "*.xlsx;*.xls")])), bootstyle="secondary-outline").grid(row=1, column=2, padx=5)
        if item:
            alias_entry.delete(0, tk.END)
            alias_entry.insert(0, item["alias"])
            alias_entry.config(foreground="black")
            alias_entry._placeholder_active = False
            path_var.set(item["path"])

        popup.grab_set()
        popup.focus_force()

        def save():
            alias = get_entry_value(alias_entry)
            path = path_var.get().strip()
            if not alias or not path:
                messagebox.showwarning("輸入不完整", "別名和路徑不能為空")
                return
            for i, ds in enumerate(self.config["data_sources"]):
                if (edit_idx is None or i != edit_idx) and ds["alias"] == alias:
                    messagebox.showwarning("別名重複", "已存在相同別名的資料來源")
                    return
            new_ds = {"alias": alias, "path": path}
            if edit_idx is not None:
                old_alias = self.config["data_sources"][edit_idx]["alias"]
                self.config["data_sources"][edit_idx] = new_ds
                for m in self.config.get("data_mappings", []):
                    if m.get("source_alias") == old_alias:
                        m["source_alias"] = alias
            else:
                self.config["data_sources"].append(new_ds)
            self.refresh_datasource_tree()
            popup.destroy()
        tb.Button(popup, text="確定", command=save, bootstyle="primary").grid(row=2, column=0, columnspan=3, pady=10)

    def delete_datasource(self):
        selected = self.ds_tree.selection()
        if not selected:
            return
        idx = self.ds_tree.index(selected[0])
        alias = self.config["data_sources"][idx]["alias"]
        refs = [m for m in self.config.get("data_mappings", []) if m.get("source_alias") == alias]
        if refs and not messagebox.askyesno("確認刪除", f"資料來源 '{alias}' 被 {len(refs)} 條對應引用，繼續？"):
            return
        del self.config["data_sources"][idx]
        self.refresh_datasource_tree()

    # ---------- 数据映射操作 ----------
    def refresh_data_tree(self):
        for i in self.data_tree.get_children():
            self.data_tree.delete(i)
        for m in self.config.get("data_mappings", []):
            self.data_tree.insert("", tk.END, values=(m.get("source_alias", ""), m["source_cell"], m["target_cell"], m.get("note", "")))

    def add_data_mapping(self, prefill=None):
        self._data_dialog(None, prefill)

    def edit_data_mapping(self):
        selected = self.data_tree.selection()
        if not selected:
            messagebox.showinfo("提示", "請先選擇一條對應")
            return
        idx = self.data_tree.index(selected[0])
        item = self.config["data_mappings"][idx]
        self._data_dialog(idx, item)

    def copy_data_mapping(self):
        selected = self.data_tree.selection()
        if not selected:
            messagebox.showinfo("提示", "請先選擇一條要複製的對應")
            return
        idx = self.data_tree.index(selected[0])
        item = self.config["data_mappings"][idx].copy()
        self._data_dialog(None, item)

    def _get_sheet_names(self, file_path):
        if not file_path or not os.path.exists(file_path):
            return []
        ext = os.path.splitext(file_path)[1].lower()
        try:
            if ext == '.xls':
                if HAS_XLRD:
                    wb = xlrd.open_workbook(file_path)
                    sheets = wb.sheet_names()
                    return sheets
            else:
                wb = load_workbook(file_path, read_only=True)
                sheets = wb.sheetnames
                wb.close()
                return sheets
        except Exception:
            return []

    def _data_dialog(self, edit_idx, item=None):
        popup = tb.Toplevel(self.root)
        popup.title("編輯資料對應" if edit_idx is not None else "新增資料對應")
        row = 0

        tb.Label(popup, text="資料來源:").grid(row=row, column=0, padx=5, pady=5, sticky=tk.W)
        alias_var = tk.StringVar()
        aliases = [ds["alias"] for ds in self.config.get("data_sources", [])]
        if not aliases:
            messagebox.showwarning("無資料來源", "請先新增資料來源")
            popup.destroy()
            return
        combo_alias = tb.Combobox(popup, textvariable=alias_var, values=aliases, state="readonly", width=28)
        combo_alias.grid(row=row, column=1, padx=5, pady=5); row += 1

        tb.Label(popup, text="來源 Sheet:").grid(row=row, column=0, padx=5, pady=5, sticky=tk.W)
        src_sheet_var = tk.StringVar()
        combo_src_sheet = tb.Combobox(popup, textvariable=src_sheet_var, width=28)
        combo_src_sheet.grid(row=row, column=1, padx=5, pady=5); row += 1

        tb.Label(popup, text="來源儲存格:").grid(row=row, column=0, padx=5, pady=5, sticky=tk.W)
        src_cell_entry = tb.Entry(popup, width=30)
        src_cell_entry.grid(row=row, column=1, padx=5, pady=5); row += 1
        add_placeholder(src_cell_entry, "A1")

        tb.Label(popup, text="目標 Sheet:").grid(row=row, column=0, padx=5, pady=5, sticky=tk.W)
        tgt_sheet_var = tk.StringVar()
        combo_tgt_sheet = tb.Combobox(popup, textvariable=tgt_sheet_var, width=28)
        combo_tgt_sheet.grid(row=row, column=1, padx=5, pady=5); row += 1

        tb.Label(popup, text="目標儲存格:").grid(row=row, column=0, padx=5, pady=5, sticky=tk.W)
        tgt_cell_entry = tb.Entry(popup, width=30)
        tgt_cell_entry.grid(row=row, column=1, padx=5, pady=5); row += 1
        add_placeholder(tgt_cell_entry, "A1")

        tb.Label(popup, text="備註:").grid(row=row, column=0, padx=5, pady=5, sticky=tk.W)
        note_entry = tb.Entry(popup, width=30)
        note_entry.grid(row=row, column=1, padx=5, pady=5); row += 1
        add_placeholder(note_entry, "Placement X")

        if item:
            alias_var.set(item.get("source_alias", aliases[0]))
            src_full = item.get("source_cell", "")
            if "!" in src_full:
                src_sheet, src_cell = src_full.split("!", 1)
            else:
                src_sheet, src_cell = "", src_full
            src_sheet_var.set(src_sheet)
            if src_cell:
                src_cell_entry.delete(0, tk.END)
                src_cell_entry.insert(0, src_cell)
                src_cell_entry.config(foreground="black")
                src_cell_entry._placeholder_active = False
            tgt_full = item.get("target_cell", "")
            if "!" in tgt_full:
                tgt_sheet, tgt_cell = tgt_full.split("!", 1)
            else:
                tgt_sheet, tgt_cell = "", tgt_full
            tgt_sheet_var.set(tgt_sheet)
            if tgt_cell:
                tgt_cell_entry.delete(0, tk.END)
                tgt_cell_entry.insert(0, tgt_cell)
                tgt_cell_entry.config(foreground="black")
                tgt_cell_entry._placeholder_active = False
            note_val = item.get("note", "")
            if note_val:
                note_entry.delete(0, tk.END)
                note_entry.insert(0, note_val)
                note_entry.config(foreground="black")
                note_entry._placeholder_active = False
        else:
            combo_alias.current(0)

        def update_src_sheets():
            alias = alias_var.get()
            ds_path = None
            for ds in self.config.get("data_sources", []):
                if ds["alias"] == alias:
                    ds_path = ds["path"]
                    break
            sheets = self._get_sheet_names(ds_path) if ds_path else []
            combo_src_sheet['values'] = sheets
            if sheets and not src_sheet_var.get():
                src_sheet_var.set(sheets[0])

        def update_tgt_sheets():
            tpl_path = self.tpl_path_var.get()
            sheets = self._get_sheet_names(tpl_path) if tpl_path else []
            combo_tgt_sheet['values'] = sheets
            if sheets and not tgt_sheet_var.get():
                tgt_sheet_var.set(sheets[0])

        combo_alias.bind("<<ComboboxSelected>>", lambda e: update_src_sheets())
        update_src_sheets()
        update_tgt_sheets()

        popup.grab_set()
        popup.focus_force()

        def save():
            alias = alias_var.get()
            src_sheet = src_sheet_var.get()
            src_cell = get_entry_value(src_cell_entry)
            tgt_sheet = tgt_sheet_var.get()
            tgt_cell = get_entry_value(tgt_cell_entry)
            note = get_entry_value(note_entry)

            if not alias or not src_sheet or not src_cell or not tgt_sheet or not tgt_cell:
                messagebox.showwarning("輸入不完整", "所有欄位不能為空")
                return

            src_full = f"{src_sheet}!{src_cell}"
            tgt_full = f"{tgt_sheet}!{tgt_cell}"

            new_map = {"source_alias": alias, "source_cell": src_full, "target_cell": tgt_full, "note": note}
            if edit_idx is not None:
                self.config["data_mappings"][edit_idx] = new_map
            else:
                self.config["data_mappings"].append(new_map)
            self.refresh_data_tree()
            popup.destroy()

        tb.Button(popup, text="確定", command=save, bootstyle="primary").grid(row=row, column=0, columnspan=2, pady=10)

    # ---------- 图片映射操作 ----------
    def refresh_image_tree(self):
        for i in self.img_tree.get_children():
            self.img_tree.delete(i)
        for m in self.config.get("image_mappings", []):
            pos_text = "預設" if m.get("position") == "top-left" else f"偏移({m.get('offset_x_cm',0)},{m.get('offset_y_cm',0)})cm"
            self.img_tree.insert("", tk.END, values=(
                m["image_number"],
                m["image_folder"],
                m["target_cell"],
                m["height_cm"],
                m["width_cm"],
                pos_text,
                m.get("note", "")
            ))

    def add_image_mapping(self, prefill=None):
        self._image_dialog(None, prefill)

    def edit_image_mapping(self):
        selected = self.img_tree.selection()
        if not selected:
            messagebox.showinfo("提示", "請先選擇一條對應")
            return
        idx = self.img_tree.index(selected[0])
        item = self.config["image_mappings"][idx]
        self._image_dialog(idx, item)

    def copy_image_mapping(self):
        selected = self.img_tree.selection()
        if not selected:
            messagebox.showinfo("提示", "請先選擇一條要複製的對應")
            return
        idx = self.img_tree.index(selected[0])
        item = self.config["image_mappings"][idx].copy()
        self._image_dialog(None, item)

    def _image_dialog(self, edit_idx, item=None):
        popup = tb.Toplevel(self.root)
        popup.title("編輯圖片對應" if edit_idx is not None else "新增圖片對應")
        row = 0

        tb.Label(popup, text="圖片編號:").grid(row=row, column=0, padx=5, pady=5, sticky=tk.W)
        num_entry = tb.Entry(popup, width=30)
        num_entry.grid(row=row, column=1, padx=5, pady=5); row += 1
        add_placeholder(num_entry, "圖片名稱")

        tb.Label(popup, text="圖片資料夾:").grid(row=row, column=0, padx=5, pady=5, sticky=tk.W)
        folder_var = tk.StringVar()
        folder_entry = tb.Entry(popup, textvariable=folder_var, width=30)
        folder_entry.grid(row=row, column=1, padx=5, pady=5)
        tb.Button(popup, text="瀏覽", command=lambda: folder_var.set(filedialog.askdirectory()), bootstyle="secondary-outline").grid(row=row, column=2, padx=5); row += 1

        tb.Label(popup, text="目標 Sheet:").grid(row=row, column=0, padx=5, pady=5, sticky=tk.W)
        tgt_sheet_var = tk.StringVar()
        combo_tgt_sheet = tb.Combobox(popup, textvariable=tgt_sheet_var, width=28)
        combo_tgt_sheet.grid(row=row, column=1, padx=5, pady=5); row += 1

        tb.Label(popup, text="目標儲存格:").grid(row=row, column=0, padx=5, pady=5, sticky=tk.W)
        tgt_cell_entry = tb.Entry(popup, width=30)
        tgt_cell_entry.grid(row=row, column=1, padx=5, pady=5); row += 1
        add_placeholder(tgt_cell_entry, "A1")

        tb.Label(popup, text="高度 (cm):").grid(row=row, column=0, padx=5, pady=5, sticky=tk.W)
        h_entry = tb.Entry(popup, width=10)
        h_entry.grid(row=row, column=1, sticky=tk.W, padx=5); row += 1

        tb.Label(popup, text="寬度 (cm):").grid(row=row, column=0, padx=5, pady=5, sticky=tk.W)
        w_entry = tb.Entry(popup, width=10)
        w_entry.grid(row=row, column=1, sticky=tk.W, padx=5); row += 1

        tb.Label(popup, text="位置:").grid(row=row, column=0, padx=5, pady=5, sticky=tk.W)
        pos_var = tk.StringVar(value="top-left")
        pos_frame = tb.Frame(popup)
        pos_frame.grid(row=row, column=1, columnspan=2, sticky=tk.W)
        tb.Radiobutton(pos_frame, text="預設（左上角）", variable=pos_var, value="top-left").pack(side=tk.LEFT)
        tb.Radiobutton(pos_frame, text="自訂偏移", variable=pos_var, value="custom").pack(side=tk.LEFT, padx=10); row += 1
        offset_frame = tb.Frame(popup)
        offset_frame.grid(row=row, column=1, columnspan=2, sticky=tk.W, pady=5)
        tb.Label(offset_frame, text="X偏移(cm):").pack(side=tk.LEFT)
        x_entry = tb.Entry(offset_frame, width=8)
        x_entry.pack(side=tk.LEFT, padx=5)
        tb.Label(offset_frame, text="Y偏移(cm):").pack(side=tk.LEFT, padx=(15,0))
        y_entry = tb.Entry(offset_frame, width=8)
        y_entry.pack(side=tk.LEFT, padx=5)
        row += 1

        def toggle_offset(*args):
            if pos_var.get() == "custom":
                x_entry.config(state="normal")
                y_entry.config(state="normal")
            else:
                x_entry.config(state="disabled")
                y_entry.config(state="disabled")
                x_entry.delete(0, tk.END); x_entry.insert(0, "0")
                y_entry.delete(0, tk.END); y_entry.insert(0, "0")
        pos_var.trace("w", toggle_offset)

        tb.Label(popup, text="備註:").grid(row=row, column=0, padx=5, pady=5, sticky=tk.W)
        note_entry = tb.Entry(popup, width=30)
        note_entry.grid(row=row, column=1, padx=5, pady=5); row += 1
        add_placeholder(note_entry, "Cosmetic")

        if item:
            num_val = item.get("image_number", "")
            if num_val:
                num_entry.delete(0, tk.END)
                num_entry.insert(0, num_val)
                num_entry.config(foreground="black")
                num_entry._placeholder_active = False
            folder_var.set(item["image_folder"])
            tgt_full = item.get("target_cell", "")
            if "!" in tgt_full:
                tgt_sheet, tgt_cell = tgt_full.split("!", 1)
            else:
                tgt_sheet, tgt_cell = "", tgt_full
            tgt_sheet_var.set(tgt_sheet)
            if tgt_cell:
                tgt_cell_entry.delete(0, tk.END)
                tgt_cell_entry.insert(0, tgt_cell)
                tgt_cell_entry.config(foreground="black")
                tgt_cell_entry._placeholder_active = False
            h_entry.insert(0, str(item["height_cm"]))
            w_entry.insert(0, str(item["width_cm"]))
            pos_var.set(item.get("position", "top-left"))
            x_entry.insert(0, str(item.get("offset_x_cm", 0)))
            y_entry.insert(0, str(item.get("offset_y_cm", 0)))
            note_val = item.get("note", "")
            if note_val:
                note_entry.delete(0, tk.END)
                note_entry.insert(0, note_val)
                note_entry.config(foreground="black")
                note_entry._placeholder_active = False
        else:
            h_entry.insert(0, "2.8"); w_entry.insert(0, "3.5")
            x_entry.insert(0, "0"); y_entry.insert(0, "0")
        toggle_offset()

        def update_tgt_sheets():
            tpl_path = self.tpl_path_var.get()
            sheets = self._get_sheet_names(tpl_path) if tpl_path else []
            combo_tgt_sheet['values'] = sheets
            if sheets and not tgt_sheet_var.get():
                tgt_sheet_var.set(sheets[0])
        update_tgt_sheets()

        popup.grab_set()
        popup.focus_force()

        def save():
            num = get_entry_value(num_entry)
            folder = folder_var.get().strip()
            tgt_sheet = tgt_sheet_var.get()
            tgt_cell = get_entry_value(tgt_cell_entry)
            try:
                h = float(h_entry.get().strip())
                w = float(w_entry.get().strip())
            except ValueError:
                messagebox.showwarning("輸入錯誤", "高度和寬度必須為數字")
                return
            if pos_var.get() == "custom":
                try:
                    off_x = float(x_entry.get().strip())
                    off_y = float(y_entry.get().strip())
                except ValueError:
                    messagebox.showwarning("輸入錯誤", "偏移值必須為數字")
                    return
            else:
                off_x = 0.0; off_y = 0.0
            note = get_entry_value(note_entry)
            if not num or not folder or not tgt_sheet or not tgt_cell:
                messagebox.showwarning("輸入不完整", "所有欄位必填")
                return
            tgt_full = f"{tgt_sheet}!{tgt_cell}"
            new_map = {
                "image_number": num,
                "image_folder": folder,
                "target_cell": tgt_full,
                "width_cm": w,
                "height_cm": h,
                "position": pos_var.get(),
                "offset_x_cm": off_x,
                "offset_y_cm": off_y,
                "note": note
            }
            if edit_idx is not None:
                self.config["image_mappings"][edit_idx] = new_map
            else:
                self.config["image_mappings"].append(new_map)
            self.refresh_image_tree()
            popup.destroy()

        tb.Button(popup, text="確定", command=save, bootstyle="primary").grid(row=row, column=0, columnspan=3, pady=10)

    def delete_selected(self, tree, map_type):
        selected = tree.selection()
        if not selected: return
        idx = tree.index(selected[0])
        if map_type == "data":
            del self.config["data_mappings"][idx]
            self.refresh_data_tree()
        else:
            del self.config["image_mappings"][idx]
            self.refresh_image_tree()

    # ---------- 配置导入导出 ----------
    def export_config(self):
        self.save_current_config()
        path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON檔案", "*.json")], initialfile="config_backup.json")
        if path:
            try:
                with open(path, "w", encoding="utf-8") as f:
                    json.dump(self.config, f, indent=2, ensure_ascii=False)
                messagebox.showinfo("匯出成功", f"設定已儲存至:\n{path}")
            except Exception as e:
                messagebox.showerror("匯出失敗", str(e))

    def import_config(self):
        path = filedialog.askopenfilename(filetypes=[("JSON檔案", "*.json")])
        if not path: return
        if not messagebox.askyesno("確認匯入", "匯入設定將覆蓋目前所有設定，確定繼續嗎？"): return
        try:
            with open(path, "r", encoding="utf-8") as f:
                new_cfg = json.load(f)
            for k in ["data_sources", "template_path", "data_mappings", "image_mappings"]:
                if k not in new_cfg:
                    raise ValueError(f"設定檔缺少必要欄位: {k}")
            self.config = new_cfg
            self.current_config_name = path
            self.tpl_path_var.set(self.config.get("template_path", ""))
            self.out_dir_var.set(self.config.get("output_dir", ""))
            self.suffix_var.set(self.config.get("output_suffix", "_已更新"))
            self.refresh_datasource_tree()
            self.refresh_data_tree()
            self.refresh_image_tree()
            messagebox.showinfo("匯入成功", "設定已匯入並更新介面")
        except Exception as e:
            messagebox.showerror("匯入失敗", f"檔案格式錯誤:\n{str(e)}")

    def save_current_config(self):
        self.config["template_path"] = self.tpl_path_var.get()
        self.config["output_dir"] = self.out_dir_var.get()
        self.config["output_suffix"] = self.suffix_var.get()
        save_config(self.config)

    # ---------- 核心执行 ----------
    def run_update(self):
        self.save_current_config()
        cfg = self.config

        if not cfg.get("template_path") or not os.path.exists(cfg["template_path"]):
            messagebox.showerror("錯誤", "範本檔案不存在")
            return

        data_wbs = {}
        for ds in cfg.get("data_sources", []):
            path = ds["path"]
            if not os.path.exists(path):
                messagebox.showerror("錯誤", f"資料來源檔案不存在: {ds['alias']} ({path})")
                return
            ext = os.path.splitext(path)[1].lower()
            try:
                if ext == '.xls':
                    if not HAS_XLRD:
                        messagebox.showerror("錯誤", "讀取 .xls 檔案需要安裝 xlrd 程式庫，請執行 pip install xlrd")
                        return
                    wb = xlrd.open_workbook(path)
                else:
                    wb = load_workbook(path, data_only=True)
                data_wbs[ds["alias"]] = wb
            except Exception as e:
                messagebox.showerror("開啟資料來源失敗", f"{ds['alias']}: {e}")
                return

        try:
            template_ext = os.path.splitext(cfg["template_path"])[1].lower()
            if template_ext == '.xls':
                messagebox.showerror("錯誤", "範本檔案目前僅支援 .xlsx 格式")
                return
            wb = load_workbook(cfg["template_path"])
        except Exception as e:
            for w in data_wbs.values():
                if isinstance(w, xlrd.Book): pass
                else: w.close()
            messagebox.showerror("開啟範本失敗", str(e))
            return

        for i, m in enumerate(cfg.get("data_mappings", [])):
            try:
                alias = m.get("source_alias")
                if alias not in data_wbs:
                    raise ValueError(f"資料來源別名 '{alias}' 不存在或未載入")
                wb_src = data_wbs[alias]
                if "!" not in m["source_cell"]: raise ValueError("缺少 '!' 分隔符")
                if "!" not in m["target_cell"]: raise ValueError("缺少 '!' 分隔符")
                src_sh, src_cell = m["source_cell"].split("!", 1)
                tgt_sh, tgt_cell = m["target_cell"].split("!", 1)

                if isinstance(wb_src, xlrd.Book):
                    if src_sh not in wb_src.sheet_names():
                        raise KeyError(f"資料來源[{alias}]中不存在工作表：{repr(src_sh)}\n可用工作表：{wb_src.sheet_names()}")
                else:
                    if src_sh not in wb_src.sheetnames:
                        raise KeyError(f"資料來源[{alias}]中不存在工作表：{repr(src_sh)}\n可用工作表：{wb_src.sheetnames}")
                if tgt_sh not in wb.sheetnames:
                    raise KeyError(f"範本中不存在工作表：{repr(tgt_sh)}\n可用工作表：{wb.sheetnames}")

                value = get_cell_value(wb_src, src_sh, src_cell)
                ws_tgt = wb[tgt_sh]
                ws_tgt[tgt_cell].value = value
            except Exception as e:
                for w in data_wbs.values():
                    if isinstance(w, xlrd.Book): pass
                    else: w.close()
                wb.close()
                messagebox.showerror("資料寫入錯誤", f"對應 {i+1}:\n來源 {m['source_cell']} → 目標 {m['target_cell']}\n錯誤：{e}")
                return

        inserted_count = 0
        skipped_details = []
        for i, m in enumerate(cfg.get("image_mappings", [])):
            try:
                number = m["image_number"]
                folder = Path(m["image_folder"])
                if not folder.exists() or not folder.is_dir():
                    skipped_details.append(f"對應{i+1}: 資料夾不存在或不是資料夾 {folder}")
                    continue
                folder_files = {}
                for f in folder.iterdir():
                    if f.is_file(): folder_files[f.name.lower()] = f.name
                if not folder_files:
                    skipped_details.append(f"對應{i+1}: 資料夾為空 {folder}")
                    continue
                img_path = None
                for ext in [".jpg", ".jpeg", ".png", ".bmp", ".gif"]:
                    candidate = f"{number}{ext}".lower()
                    if candidate in folder_files:
                        img_path = str(folder / folder_files[candidate])
                        break
                if not img_path:
                    file_list = "\n".join(sorted(folder_files.values())[:15])
                    skipped_details.append(f"對應{i+1}: 未找到編號 '{number}' 的圖片\n資料夾內容(前15):\n{file_list}")
                    continue
                if "!" not in m["target_cell"]:
                    skipped_details.append(f"對應{i+1}: 目標儲存格格式錯誤")
                    continue
                tgt_sh, tgt_cell = m["target_cell"].split("!", 1)

                if tgt_sh not in wb.sheetnames:
                    skipped_details.append(f"對應{i+1}: 範本中不存在工作表 {repr(tgt_sh)}")
                    continue

                ws_tgt = wb[tgt_sh]
                img = XLImage(img_path)
                img.width = cm_to_px(m["width_cm"])
                img.height = cm_to_px(m["height_cm"])
                ws_tgt.add_image(img, tgt_cell)

                if m.get("position") == "custom":
                    off_x = m.get("offset_x_cm", 0)
                    off_y = m.get("offset_y_cm", 0)
                    anchor = img.anchor
                    if isinstance(anchor, (OneCellAnchor, TwoCellAnchor)):
                        from_marker = anchor._from
                        from_marker.colOff = cm_to_emu(off_x)
                        from_marker.rowOff = cm_to_emu(off_y)
                    elif isinstance(anchor, str):
                        if HAS_XDR:
                            row_num, col_idx = coordinate_to_tuple(anchor)
                            marker = AnchorMarker(col=col_idx - 1, row=row_num - 1,
                                                  colOff=cm_to_emu(off_x), rowOff=cm_to_emu(off_y))
                            ext = XDRPositiveSize2D(cx=cm_to_emu(m["width_cm"]), cy=cm_to_emu(m["height_cm"]))
                            img.anchor = OneCellAnchor(_from=marker, ext=ext)
                        else:
                            skipped_details.append(f"對應{i+1}: 自訂偏移未生效（缺少XDR支援）")
                    else:
                        skipped_details.append(f"對應{i+1}: 無法設定偏移，未知錨點類型")
                inserted_count += 1
            except Exception as e:
                for w in data_wbs.values():
                    if isinstance(w, xlrd.Book): pass
                    else: w.close()
                wb.close()
                messagebox.showerror("圖片插入錯誤", f"圖片對應 {i+1}:\n編號 {m['image_number']}，目標 {m['target_cell']}\n錯誤：{e}")
                return

        for w in data_wbs.values():
            if isinstance(w, xlrd.Book): pass
            else: w.close()

        tpl_path = Path(cfg["template_path"])
        suffix = cfg.get("output_suffix", "_已更新")
        out_dir = cfg.get("output_dir", "")
        if out_dir and os.path.isdir(out_dir):
            out_path = Path(out_dir) / f"{tpl_path.stem}{suffix}.xlsx"
        else:
            out_path = tpl_path.parent / f"{tpl_path.stem}{suffix}.xlsx"
        try:
            wb.save(str(out_path))
        except Exception as e:
            wb.close()
            messagebox.showerror("儲存失敗", str(e))
            return
        wb.close()

        summary = f"資料來源: {len(data_wbs)} 個\n資料對應: {len(cfg.get('data_mappings', []))} 條\n圖片對應: {len(cfg.get('image_mappings', []))} 條\n成功插入圖片: {inserted_count} 張\n"
        if skipped_details:
            summary += "\n未插入圖片原因:\n" + "\n\n".join(skipped_details)
        messagebox.showinfo("執行結果", summary)
        try:
            os.startfile(out_path)
        except Exception:
            pass

if __name__ == "__main__":
    root = tb.Window(themename="flatly")
    app = App(root)
    root.update_idletasks()
    req_width = root.winfo_reqwidth()
    req_height = root.winfo_reqheight()
    root.geometry(f"{req_width}x{req_height}")
    root.minsize(req_width, req_height)
    root.resizable(True, True)
    root.mainloop()
