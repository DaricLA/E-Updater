import tkinter as tk
from tkinter import ttk, filedialog, messagebox, colorchooser
import ttkbootstrap as tb
from ttkbootstrap.constants import *
import xml.etree.ElementTree as ET
import os
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment

# ====================== 预定义颜色 ======================
PREDEFINED_COLORS = [
    ("白色", "#FFFFFF"), ("黑色", "#000000"), ("红色", "#FF0000"),
    ("綠色", "#00AA00"), ("藍色", "#0000FF"), ("黃色", "#FFFF00"),
    ("青色", "#00FFFF"), ("洋紅", "#FF00FF"), ("灰色", "#AAAAAA"),
    ("橙色", "#FF8000"), ("紫色", "#800080"), ("棕色", "#8B4513"),
    ("粉色", "#FFC0CB"), ("淺綠", "#90EE90"), ("淺藍", "#ADD8E6"),
]

FONT = ("微軟雅黑", 9)
FONT_BOLD = ("微軟雅黑", 9, "bold")
FONT_MATRIX = ("微軟雅黑", 8)

class WaferMapApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Wafer Map 轉換工具")
        self.root.geometry("1200x800")
        self.root.resizable(True, True)

        self.xml_path = ""
        self.wafer_info_text = ""
        self.matrix = []
        self.rows = 0
        self.cols = 0
        self.transformed_data = []
        self.bin_colors = {}
        self.unique_bins = []
        self.color_rules = []       # 第一类规则：[{"bin":, "label":, "color":}]
        self.special_markers = []   # 第二类标记：[{"x":, "y":, "bin":, "label":, "color":}]

        self.zoom_scale = 1.0
        self.min_zoom = 0.3
        self.max_zoom = 2.0          # 缩放上限改为200%
        self.zoom_step = 0.05
        self.base_cell_height = 12
        self.base_cell_width = int(12 * 1.1)

        # 页面容器
        self.page1 = tb.Frame(self.root)
        self.page2 = tb.Frame(self.root)

        self.build_page1()
        self.build_page2()
        self.page1.pack(fill=BOTH, expand=YES)

    # ==================== 第一页构建 ====================
    def build_page1(self):
        top_bar = tb.Frame(self.page1, padding=10)
        top_bar.pack(fill=X)
        tb.Button(top_bar, text="選擇 XML 文件", command=self.load_xml,
                  bootstyle="primary", width=14).pack(side=LEFT)
        self.file_label = tb.Label(top_bar, text="尚未選擇文件",
                                   foreground="gray", font=FONT)
        self.file_label.pack(side=LEFT, padx=10)

        main_pw = tk.PanedWindow(self.page1, orient=tk.HORIZONTAL,
                                 sashrelief=tk.RAISED, sashwidth=4)
        main_pw.pack(fill=BOTH, expand=YES, padx=10, pady=(0, 5))

        # ----- 左侧面板 -----
        left_frame = tb.Frame(main_pw, width=400)
        left_frame.pack_propagate(False)

        # 表头信息
        info_lf = tb.LabelFrame(left_frame, text="表頭完整信息", padding=5, bootstyle="info")
        info_lf.pack(fill=BOTH, expand=YES, padx=2, pady=(0, 3))

        self.info_text = tk.Text(info_lf, wrap="none", state="disabled",
                                 font=FONT, relief=tk.FLAT)
        v_scroll = tb.Scrollbar(info_lf, orient=VERTICAL, command=self.info_text.yview, bootstyle="round")
        h_scroll = tb.Scrollbar(info_lf, orient=HORIZONTAL, command=self.info_text.xview, bootstyle="round")
        self.info_text.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)
        self.info_text.grid(row=0, column=0, sticky=NSEW)
        v_scroll.grid(row=0, column=1, sticky=NS)
        h_scroll.grid(row=1, column=0, sticky=EW)
        info_lf.grid_rowconfigure(0, weight=1)
        info_lf.grid_columnconfigure(0, weight=1)

        # 标签颜色规则（第一类）
        rule_lf = tb.LabelFrame(left_frame, text="標籤顏色規則 (Bin+Label)", padding=5, bootstyle="info")
        rule_lf.pack(fill=BOTH, expand=YES, padx=2, pady=2)

        rule_container = tb.Frame(rule_lf)
        rule_container.pack(fill=BOTH, expand=YES)
        self.rule_tree = tb.Treeview(rule_container, columns=("bin", "label", "color"), show="headings", height=5, bootstyle="primary")
        self.rule_tree.heading("bin", text="Bin")
        self.rule_tree.heading("label", text="Label")
        self.rule_tree.heading("color", text="顏色")
        self.rule_tree.column("bin", width=50, anchor=CENTER)
        self.rule_tree.column("label", width=120, anchor=CENTER)
        self.rule_tree.column("color", width=100, anchor=CENTER)
        vscroll_rule = tb.Scrollbar(rule_container, orient=VERTICAL, command=self.rule_tree.yview, bootstyle="round")
        self.rule_tree.configure(yscrollcommand=vscroll_rule.set)
        self.rule_tree.grid(row=0, column=0, sticky=NSEW)
        vscroll_rule.grid(row=0, column=1, sticky=NS)
        rule_container.grid_rowconfigure(0, weight=1)
        rule_container.grid_columnconfigure(0, weight=1)

        btn_rule = tb.Frame(rule_lf)
        btn_rule.pack(fill=X, pady=2)
        tb.Button(btn_rule, text="新增規則", command=self.add_color_rule, bootstyle="primary-outline").pack(side=LEFT, padx=2)
        tb.Button(btn_rule, text="編輯選中", command=self.edit_color_rule, bootstyle="secondary-outline").pack(side=LEFT, padx=2)
        tb.Button(btn_rule, text="刪除選中", command=self.delete_color_rule, bootstyle="danger-outline").pack(side=LEFT, padx=2)

        # 特殊坐标标记（第二类）
        marker_lf = tb.LabelFrame(left_frame, text="特殊座標標記", padding=5, bootstyle="info")
        marker_lf.pack(fill=BOTH, expand=YES, padx=2, pady=2)

        marker_container = tb.Frame(marker_lf)
        marker_container.pack(fill=BOTH, expand=YES)
        self.marker_tree = tb.Treeview(marker_container, columns=("x", "y", "bin", "label", "color"), show="headings", height=5, bootstyle="primary")
        self.marker_tree.heading("x", text="X")
        self.marker_tree.heading("y", text="Y")
        self.marker_tree.heading("bin", text="Bin")
        self.marker_tree.heading("label", text="Label")
        self.marker_tree.heading("color", text="顏色")
        self.marker_tree.column("x", width=40, anchor=CENTER)
        self.marker_tree.column("y", width=40, anchor=CENTER)
        self.marker_tree.column("bin", width=50, anchor=CENTER)
        self.marker_tree.column("label", width=100, anchor=CENTER)
        self.marker_tree.column("color", width=80, anchor=CENTER)
        vscroll_marker = tb.Scrollbar(marker_container, orient=VERTICAL, command=self.marker_tree.yview, bootstyle="round")
        self.marker_tree.configure(yscrollcommand=vscroll_marker.set)
        self.marker_tree.grid(row=0, column=0, sticky=NSEW)
        vscroll_marker.grid(row=0, column=1, sticky=NS)
        marker_container.grid_rowconfigure(0, weight=1)
        marker_container.grid_columnconfigure(0, weight=1)

        btn_marker = tb.Frame(marker_lf)
        btn_marker.pack(fill=X, pady=2)
        tb.Button(btn_marker, text="新增標記", command=self.add_special_marker, bootstyle="primary-outline").pack(side=LEFT, padx=2)
        tb.Button(btn_marker, text="編輯選中", command=self.edit_special_marker, bootstyle="secondary-outline").pack(side=LEFT, padx=2)
        tb.Button(btn_marker, text="刪除選中", command=self.delete_special_marker, bootstyle="danger-outline").pack(side=LEFT, padx=2)

        main_pw.add(left_frame, width=400)

        # ----- 右侧面板 -----
        right_frame = tb.Frame(main_pw)
        matrix_lf = tb.LabelFrame(right_frame, text="Wafer 矩陣預覽", padding=5, bootstyle="info")
        matrix_lf.pack(fill=BOTH, expand=YES)

        zoom_bar = tb.Frame(matrix_lf)
        zoom_bar.pack(fill=X, pady=2)
        tb.Button(zoom_bar, text="−", command=self.zoom_out, bootstyle="secondary", width=3).pack(side=LEFT, padx=2)
        self.zoom_scale_var = tk.DoubleVar(value=1.0)
        self.zoom_slider = tb.Scale(zoom_bar, from_=self.min_zoom, to=self.max_zoom,
                                    orient=HORIZONTAL, variable=self.zoom_scale_var,
                                    command=self.on_zoom_slider, bootstyle="primary")
        self.zoom_slider.pack(side=LEFT, fill=X, expand=YES, padx=5)
        self.zoom_slider.set(1.0)
        tb.Button(zoom_bar, text="+", command=self.zoom_in, bootstyle="secondary", width=3).pack(side=LEFT, padx=2)
        self.zoom_label = tb.Label(zoom_bar, text="100%", font=FONT, width=5)
        self.zoom_label.pack(side=LEFT, padx=5)
        tb.Button(zoom_bar, text="複製矩陣文字", command=self.copy_matrix_text, bootstyle="info", width=12).pack(side=RIGHT, padx=5)

        canvas_frame = tb.Frame(matrix_lf)
        canvas_frame.pack(fill=BOTH, expand=YES)
        self.canvas = tk.Canvas(canvas_frame, bg="white", cursor="arrow")
        h_scroll_canvas = tb.Scrollbar(canvas_frame, orient=HORIZONTAL, command=self.canvas.xview, bootstyle="round")
        v_scroll_canvas = tb.Scrollbar(canvas_frame, orient=VERTICAL, command=self.canvas.yview, bootstyle="round")
        self.canvas.configure(xscrollcommand=h_scroll_canvas.set, yscrollcommand=v_scroll_canvas.set)
        self.canvas.grid(row=0, column=0, sticky=NSEW)
        h_scroll_canvas.grid(row=1, column=0, sticky=EW)
        v_scroll_canvas.grid(row=0, column=1, sticky=NS)
        canvas_frame.grid_rowconfigure(0, weight=1)
        canvas_frame.grid_columnconfigure(0, weight=1)

        main_pw.add(right_frame)

        # 底部转换按钮
        tb.Button(self.page1, text="轉換 → 生成坐標列表", command=self.go_to_page2,
                  bootstyle="primary", padding=(20, 10), width=25).pack(pady=10)

    # ==================== 第二页构建 ====================
    def build_page2(self):
        top_frame = tb.Frame(self.page2, padding=10)
        top_frame.pack(fill=X)
        tb.Button(top_frame, text="← 返回預覽", command=self.go_to_page1, bootstyle="secondary", width=12).pack(side=LEFT)
        self.status_label = tb.Label(top_frame, text="", foreground="blue", font=FONT)
        self.status_label.pack(side=LEFT, padx=20)

        table_frame = tb.Frame(self.page2, padding=10)
        table_frame.pack(fill=BOTH, expand=YES)

        columns = ("RW_XY", "Bin", "X", "Y", "Label")
        self.tree = tb.Treeview(table_frame, columns=columns, show="headings", height=25, bootstyle="primary")
        self.tree.heading("RW_XY", text="RW X_Y")
        self.tree.heading("Bin", text="Bin")
        self.tree.heading("X", text="X")
        self.tree.heading("Y", text="Y")
        self.tree.heading("Label", text="Label")
        self.tree.column("RW_XY", width=100, anchor=CENTER)
        self.tree.column("Bin", width=60, anchor=CENTER)
        self.tree.column("X", width=80, anchor=CENTER)
        self.tree.column("Y", width=80, anchor=CENTER)
        self.tree.column("Label", width=120, anchor=CENTER)

        scroll_y = tb.Scrollbar(table_frame, orient=VERTICAL, command=self.tree.yview, bootstyle="round")
        scroll_x = tb.Scrollbar(table_frame, orient=HORIZONTAL, command=self.tree.xview, bootstyle="round")
        self.tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        self.tree.grid(row=0, column=0, sticky=NSEW)
        scroll_y.grid(row=0, column=1, sticky=NS)
        scroll_x.grid(row=1, column=0, sticky=EW)
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        self.tree.bind("<Control-c>", self.copy_tree_selection)
        self.tree.bind("<Control-C>", self.copy_tree_selection)

        btn_frame = tb.Frame(self.page2, padding=10)
        btn_frame.pack()
        tb.Button(btn_frame, text="複製全部數據到剪貼板", command=self.copy_all, bootstyle="primary", width=25).pack(side=LEFT, padx=5)
        tb.Button(btn_frame, text="導出 Excel", command=self.export_excel, bootstyle="success", width=15).pack(side=LEFT, padx=5)

    # ==================== 页面切换 ====================
    def go_to_page1(self):
        self.page2.pack_forget()
        self.page1.pack(fill=BOTH, expand=YES)

    def go_to_page2(self):
        if not self.matrix:
            messagebox.showwarning("警告", "請先載入有效的 XML 文件。")
            return
        self.transform()
        self.page1.pack_forget()
        self.page2.pack(fill=BOTH, expand=YES)

    # ==================== XML 加载 ====================
    def load_xml(self):
        path = filedialog.askopenfilename(title="選擇 Wafer Map XML 文件",
                                          filetypes=[("XML 文件", "*.xml"), ("所有文件", "*.*")])
        if not path:
            return

        try:
            tree = ET.parse(path)
            root = tree.getroot()
        except Exception as e:
            messagebox.showerror("錯誤", f"無法解析 XML 文件:\n{e}")
            return

        device = root.find(".//{*}Device")
        if device is None:
            messagebox.showerror("錯誤", "找不到 <Device> 節點")
            return

        self.wafer_info_text = self.build_header_text(root, device)

        cols_str = device.get("Columns", "0")
        rows_str = device.get("Rows", "0")
        try:
            self.cols = int(cols_str)
            self.rows = int(rows_str)
        except:
            messagebox.showerror("錯誤", "Columns 或 Rows 屬性無效")
            return

        null_bin = device.get("NullBin", "F")
        data = device.find(".//{*}Data")
        if data is None:
            messagebox.showerror("錯誤", "找不到 <Data> 節點")
            return

        rows_elem = data.findall(".//{*}Row")
        raw_lines = []
        for row in rows_elem:
            text = row.text
            if text:
                text = text.strip()
                if text:
                    raw_lines.append(text)

        self.matrix = []
        for line in raw_lines:
            if len(line) < self.cols:
                line = line.ljust(self.cols, null_bin)
            else:
                line = line[:self.cols]
            self.matrix.append(list(line))

        while len(self.matrix) < self.rows:
            self.matrix.append([null_bin] * self.cols)

        all_chars = set()
        for row in self.matrix:
            all_chars.update(row)
        self.unique_bins = sorted(list(all_chars))

        self.auto_color_map(device, data)
        self.update_header_display()
        self.zoom_scale = 1.0
        self.zoom_scale_var.set(1.0)
        self.xml_path = path
        self.file_label.config(text=os.path.basename(path), foreground="black")
        self.draw_matrix()

    def build_header_text(self, root, device):
        lines = []
        map_elem = root.find(".//{*}Map")
        if map_elem is None:
            map_elem = root
        lines.append("[Map]")
        for k, v in map_elem.attrib.items():
            lines.append(f"  {k} = {v}")
        lines.append("\n[Device]")
        for k, v in device.attrib.items():
            lines.append(f"  {k} = {v}")
        ref = device.find(".//{*}ReferenceDevice")
        if ref is not None:
            lines.append("\n[ReferenceDevice]")
            for k, v in ref.attrib.items():
                lines.append(f"  {k} = {v}")
        data = device.find(".//{*}Data")
        if data is not None:
            lines.append("\n[Data]")
            for k, v in data.attrib.items():
                lines.append(f"  {k} = {v}")
            bins = data.findall(".//{*}Bin")
            if bins:
                lines.append("\n[Bin Definitions]")
                for i, bin_elem in enumerate(bins, 1):
                    attrs = ", ".join([f"{k}={v}" for k, v in bin_elem.attrib.items()])
                    lines.append(f"  Bin{i}: {attrs}")
        return "\n".join(lines)

    def auto_color_map(self, device, data):
        self.bin_colors.clear()
        bins = data.findall(".//{*}Bin") if data else []
        if bins:
            for bin_elem in bins:
                code = bin_elem.get("BinCode", "")
                quality = bin_elem.get("BinQuality", "")
                if code:
                    if "pass" in quality.lower():
                        color = "#90EE90"
                    elif "fail" in quality.lower() or "reject" in quality.lower():
                        color = "#FF9999"
                    else:
                        color = "#FFFFFF"
                    self.bin_colors[code] = color
        else:
            self.bin_colors = {
                'F': '#CCCCCC', 'y': '#90EE90',
                'x': '#FF9999', 'X': '#8B4513'
            }
        for code in self.unique_bins:
            if code not in self.bin_colors:
                self.bin_colors[code] = "#FFFFFF"

    # ==================== 界面更新 ====================
    def update_header_display(self):
        self.info_text.config(state="normal")
        self.info_text.delete("1.0", "end")
        self.info_text.insert("1.0", self.wafer_info_text)
        self.info_text.config(state="disabled")

    # ==================== 矩阵绘制 ====================
    def get_cell_size(self):
        w = int(self.base_cell_width * self.zoom_scale)
        h = int(self.base_cell_height * self.zoom_scale)
        w = max(w, 2)
        h = max(h, 2)
        return w, h

    def get_cell_label(self, bin_char, x, y):
        """获取坐标对应的标签，优先特殊标记，其次第一类规则中第一个匹配的Label"""
        # 特殊标记优先
        for marker in self.special_markers:
            if marker["x"] == x and marker["y"] == y:
                return marker["label"]
        # 否则从规则中找第一个匹配 Bin 的 Label
        for rule in self.color_rules:
            if rule["bin"] == bin_char:
                return rule["label"]
        return ""

    def get_cell_color(self, bin_char, x, y):
        """根据规则返回最终单元格颜色"""
        # 特殊标记优先
        for marker in self.special_markers:
            if marker["x"] == x and marker["y"] == y:
                return marker["color"]
        # 从规则中查找匹配的 Bin+Label
        label = self.get_cell_label(bin_char, x, y)
        if label:
            for rule in self.color_rules:
                if rule["bin"] == bin_char and rule["label"] == label:
                    return rule["color"]
        # 否则使用默认颜色
        return self.bin_colors.get(bin_char, "#FFFFFF")

    def draw_matrix(self):
        self.canvas.delete("all")
        if not self.matrix:
            return

        cell_w, cell_h = self.get_cell_size()
        rows = self.rows
        cols = self.cols
        total_w = cols * cell_w
        total_h = rows * cell_h
        self.canvas.config(scrollregion=(0, 0, total_w, total_h))
        font_size = max(6, int(min(cell_w, cell_h) * 0.65))

        for r in range(rows):
            for c in range(cols):
                x1 = c * cell_w
                y1 = r * cell_h
                x2 = x1 + cell_w
                y2 = y1 + cell_h
                char = self.matrix[r][c]
                color = self.get_cell_color(char, c+1, r+1)
                self.canvas.create_rectangle(x1, y1, x2, y2,
                                             fill=color, outline="#D0D0D0", width=1)
                self.canvas.create_text(x1 + cell_w/2, y1 + cell_h/2,
                                        text=char, font=("微軟雅黑", font_size),
                                        fill="black")
        self.update_zoom_label()

    def update_zoom_label(self):
        percent = int(self.zoom_scale * 100)
        self.zoom_label.config(text=f"{percent}%")
        self.zoom_scale_var.set(self.zoom_scale)

    def zoom_in(self):
        new_scale = round(self.zoom_scale + self.zoom_step, 2)
        if new_scale <= self.max_zoom:
            self.zoom_scale = new_scale
            self.draw_matrix()

    def zoom_out(self):
        new_scale = round(self.zoom_scale - self.zoom_step, 2)
        if new_scale >= self.min_zoom:
            self.zoom_scale = new_scale
            self.draw_matrix()

    def on_zoom_slider(self, event=None):
        raw = self.zoom_scale_var.get()
        steps = round(raw / self.zoom_step)
        val = steps * self.zoom_step
        val = max(self.min_zoom, min(self.max_zoom, val))
        if val != self.zoom_scale:
            self.zoom_scale = val
            self.draw_matrix()
        else:
            self.zoom_scale_var.set(self.zoom_scale)

    # ==================== 颜色规则管理 ====================
    def refresh_rule_tree(self):
        for i in self.rule_tree.get_children():
            self.rule_tree.delete(i)
        for rule in self.color_rules:
            self.rule_tree.insert("", "end", values=(rule["bin"], rule["label"], rule["color"]))

    def refresh_marker_tree(self):
        for i in self.marker_tree.get_children():
            self.marker_tree.delete(i)
        for marker in self.special_markers:
            self.marker_tree.insert("", "end", values=(marker["x"], marker["y"], marker["bin"], marker["label"], marker["color"]))

    def choose_color(self, initial_color="#FFFFFF"):
        color_code = colorchooser.askcolor(initialcolor=initial_color, title="選擇顏色")
        if color_code and color_code[1]:
            return color_code[1]
        return initial_color

    def add_color_rule(self):
        if not self.unique_bins:
            messagebox.showinfo("提示", "請先載入矩陣。")
            return
        self._color_rule_dialog(None)

    def edit_color_rule(self):
        selected = self.rule_tree.selection()
        if not selected:
            messagebox.showinfo("提示", "請先選擇一條規則")
            return
        idx = self.rule_tree.index(selected[0])
        rule = self.color_rules[idx]
        self._color_rule_dialog(idx, rule)

    def _color_rule_dialog(self, edit_idx=None, rule=None):
        dialog = tb.Toplevel(self.root)
        dialog.title("編輯標籤顏色規則" if edit_idx is not None else "新增標籤顏色規則")
        row = 0
        tb.Label(dialog, text="Bin:").grid(row=row, column=0, padx=5, pady=5, sticky=tk.W)
        bin_var = tk.StringVar()
        combo = tb.Combobox(dialog, textvariable=bin_var, values=self.unique_bins, state="readonly", width=10)
        combo.grid(row=row, column=1, padx=5, pady=5)
        if rule:
            bin_var.set(rule["bin"])
        elif self.unique_bins:
            combo.current(0)
        row += 1
        tb.Label(dialog, text="Label:").grid(row=row, column=0, padx=5, pady=5, sticky=tk.W)
        label_entry = tb.Entry(dialog, width=30)
        label_entry.grid(row=row, column=1, padx=5, pady=5)
        if rule:
            label_entry.insert(0, rule["label"])
        row += 1
        tb.Label(dialog, text="顏色:").grid(row=row, column=0, padx=5, pady=5, sticky=tk.W)
        color_var = tk.StringVar(value=rule["color"] if rule else "#FFFFFF")
        color_preview = tb.Label(dialog, text="   ", background=color_var.get(), relief="ridge", width=5)
        color_preview.grid(row=row, column=1, sticky=tk.W, padx=5)
        tb.Button(dialog, text="選擇", command=lambda: self._pick_color_dialog(color_var, color_preview), bootstyle="secondary-outline").grid(row=row, column=2, padx=5)
        row += 1

        def save_rule():
            bin_code = bin_var.get()
            label = label_entry.get().strip()
            color = color_var.get()
            if not label:
                messagebox.showwarning("輸入不完整", "Label 不能為空")
                return
            # 检查重复
            for i, r in enumerate(self.color_rules):
                if (edit_idx is None or i != edit_idx) and r["bin"] == bin_code and r["label"] == label:
                    messagebox.showerror("重複", "同一 Bin + Label 只能有一條規則")
                    return
            new_rule = {"bin": bin_code, "label": label, "color": color}
            if edit_idx is not None:
                self.color_rules[edit_idx] = new_rule
            else:
                self.color_rules.append(new_rule)
            self.refresh_rule_tree()
            self.draw_matrix()
            dialog.destroy()

        tb.Button(dialog, text="確定", command=save_rule, bootstyle="primary").grid(row=row, column=0, columnspan=3, pady=10)
        dialog.grab_set()
        dialog.focus_force()

    def _pick_color_dialog(self, color_var, preview_label):
        color = colorchooser.askcolor(initialcolor=color_var.get(), title="選擇顏色")
        if color and color[1]:
            color_var.set(color[1])
            preview_label.config(background=color[1])

    def delete_color_rule(self):
        selected = self.rule_tree.selection()
        if not selected:
            return
        idx = self.rule_tree.index(selected[0])
        del self.color_rules[idx]
        self.refresh_rule_tree()
        self.draw_matrix()

    # 特殊标记
    def add_special_marker(self):
        self._special_marker_dialog(None)

    def edit_special_marker(self):
        selected = self.marker_tree.selection()
        if not selected:
            messagebox.showinfo("提示", "請先選擇一個標記")
            return
        idx = self.marker_tree.index(selected[0])
        marker = self.special_markers[idx]
        self._special_marker_dialog(idx, marker)

    def _special_marker_dialog(self, edit_idx=None, marker=None):
        dialog = tb.Toplevel(self.root)
        dialog.title("編輯特殊座標標記" if edit_idx is not None else "新增特殊座標標記")
        row = 0
        tb.Label(dialog, text="X:").grid(row=row, column=0, padx=5, pady=5, sticky=tk.W)
        x_entry = tb.Entry(dialog, width=8)
        x_entry.grid(row=row, column=1, padx=5, pady=5)
        if marker:
            x_entry.insert(0, str(marker["x"]))
        row += 1
        tb.Label(dialog, text="Y:").grid(row=row, column=0, padx=5, pady=5, sticky=tk.W)
        y_entry = tb.Entry(dialog, width=8)
        y_entry.grid(row=row, column=1, padx=5, pady=5)
        if marker:
            y_entry.insert(0, str(marker["y"]))
        row += 1
        tb.Label(dialog, text="Bin:").grid(row=row, column=0, padx=5, pady=5, sticky=tk.W)
        bin_var = tk.StringVar()
        bin_combo = tb.Combobox(dialog, textvariable=bin_var, values=self.unique_bins, state="readonly", width=10)
        bin_combo.grid(row=row, column=1, padx=5, pady=5)
        if marker:
            bin_var.set(marker["bin"])
        row += 1

        def auto_bin():
            try:
                x = int(x_entry.get().strip())
                y = int(y_entry.get().strip())
            except ValueError:
                messagebox.showwarning("輸入錯誤", "X 和 Y 必須為整數")
                return
            if 1 <= x <= self.cols and 1 <= y <= self.rows:
                bin_code = self.matrix[y-1][x-1]
                bin_var.set(bin_code)
            else:
                messagebox.showwarning("範圍錯誤", "X 或 Y 超出矩陣範圍")

        tb.Button(dialog, text="自動帶出", command=auto_bin, bootstyle="secondary-outline").grid(row=row, column=2, padx=5)
        row += 1
        tb.Label(dialog, text="Label:").grid(row=row, column=0, padx=5, pady=5, sticky=tk.W)
        label_entry = tb.Entry(dialog, width=30)
        label_entry.grid(row=row, column=1, padx=5, pady=5)
        if marker:
            label_entry.insert(0, marker["label"])
        row += 1
        tb.Label(dialog, text="顏色:").grid(row=row, column=0, padx=5, pady=5, sticky=tk.W)
        color_var = tk.StringVar(value=marker["color"] if marker else "#FFFFFF")
        color_preview = tb.Label(dialog, text="   ", background=color_var.get(), relief="ridge", width=5)
        color_preview.grid(row=row, column=1, sticky=tk.W, padx=5)
        tb.Button(dialog, text="選擇", command=lambda: self._pick_color_dialog(color_var, color_preview), bootstyle="secondary-outline").grid(row=row, column=2, padx=5)
        row += 1

        def save_marker():
            try:
                x = int(x_entry.get().strip())
                y = int(y_entry.get().strip())
            except ValueError:
                messagebox.showwarning("輸入錯誤", "X 和 Y 必須為整數")
                return
            if not (1 <= x <= self.cols and 1 <= y <= self.rows):
                messagebox.showwarning("範圍錯誤", "X 或 Y 超出矩陣範圍")
                return
            bin_code = bin_var.get()
            label = label_entry.get().strip()
            color = color_var.get()
            if not label:
                messagebox.showwarning("輸入不完整", "Label 不能為空")
                return
            # 检查与第一类规则重复
            for rule in self.color_rules:
                if rule["bin"] == bin_code and rule["label"] == label:
                    messagebox.showerror("重複", "與標籤顏色規則重複，無法保存")
                    return
            # 检查标记重复
            for i, m in enumerate(self.special_markers):
                if (edit_idx is None or i != edit_idx) and m["x"] == x and m["y"] == y:
                    messagebox.showerror("重複", "該坐標已存在標記")
                    return
            new_marker = {"x": x, "y": y, "bin": bin_code, "label": label, "color": color}
            if edit_idx is not None:
                self.special_markers[edit_idx] = new_marker
            else:
                self.special_markers.append(new_marker)
            self.refresh_marker_tree()
            self.draw_matrix()
            dialog.destroy()

        tb.Button(dialog, text="確定", command=save_marker, bootstyle="primary").grid(row=row, column=0, columnspan=3, pady=10)
        dialog.grab_set()
        dialog.focus_force()

    def delete_special_marker(self):
        selected = self.marker_tree.selection()
        if not selected:
            return
        idx = self.marker_tree.index(selected[0])
        del self.special_markers[idx]
        self.refresh_marker_tree()
        self.draw_matrix()

    # ==================== 转换结果 ====================
    def transform(self):
        self.transformed_data.clear()
        for item in self.tree.get_children():
            self.tree.delete(item)

        for y in range(1, self.rows + 1):
            for x in range(1, self.cols + 1):
                bin_char = self.matrix[y-1][x-1]
                rw_xy = f"{x}_{y}"
                label = self.get_cell_label(bin_char, x, y)
                self.transformed_data.append((rw_xy, bin_char, x, y, label))
                self.tree.insert("", "end", values=(rw_xy, bin_char, x, y, label))

        self.status_label.config(text=f"共 {len(self.transformed_data)} 條記錄")

    # ==================== 复制功能 ====================
    def copy_tree_selection(self, event=None):
        selected = self.tree.selection()
        if not selected:
            return
        lines = []
        for item in selected:
            values = self.tree.item(item, "values")
            lines.append("\t".join(str(v) for v in values))
        self.root.clipboard_clear()
        self.root.clipboard_append("\n".join(lines))

    def copy_all(self):
        if not self.transformed_data:
            return
        lines = ["RW X_Y\tBin\tX\tY\tLabel"]
        for item in self.transformed_data:
            lines.append("\t".join(str(v) for v in item))
        text = "\n".join(lines)
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        messagebox.showinfo("成功", "數據已複製到剪貼板。")

    # ==================== 导出 Excel ====================
    def export_excel(self):
        if not self.transformed_data:
            messagebox.showwarning("警告", "請先進行轉換生成坐標列表")
            return
        if not self.xml_path:
            messagebox.showwarning("警告", "請先載入 XML 文件")
            return

        base_path = os.path.splitext(self.xml_path)[0]
        output_path = base_path + "_output.xlsx"

        wb = Workbook()
        ws_coord = wb.active
        ws_coord.title = "Coordinates"
        headers = ["RW X_Y", "Bin", "X", "Y", "Label"]
        ws_coord.append(headers)
        for item in self.transformed_data:
            ws_coord.append(list(item))

        ws_map = wb.create_sheet("ColorMap")
        for r in range(self.rows):
            row_data = []
            for c in range(self.cols):
                row_data.append(self.matrix[r][c])
            ws_map.append(row_data)

        fill_cache = {}
        for r in range(self.rows):
            for c in range(self.cols):
                cell = ws_map.cell(row=r+1, column=c+1)
                char = self.matrix[r][c]
                color = self.get_cell_color(char, c+1, r+1)
                if color not in fill_cache:
                    fill_cache[color] = PatternFill(start_color=color[1:], end_color=color[1:], fill_type="solid")
                cell.fill = fill_cache[color]
                cell.alignment = Alignment(horizontal='center', vertical='center')

        try:
            wb.save(output_path)
            messagebox.showinfo("成功", f"Excel 已保存至:\n{output_path}")
            os.startfile(output_path)
        except Exception as e:
            messagebox.showerror("錯誤", f"保存失敗:\n{e}")

if __name__ == "__main__":
    root = tb.Window(themename="flatly")
    app = WaferMapApp(root)
    root.mainloop()
